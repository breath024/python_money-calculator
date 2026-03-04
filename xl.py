import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from datetime import datetime, timedelta

style = NamedStyle(name="currency_krw", number_format='₩#,##0')
minus_style = NamedStyle(name="currency_krw_minus", number_format='₩#,##0;[Red]-₩#,##0')

dlt = datetime(datetime.now().year, 1, 1)
print(dlt)
try:
    wb = load_workbook(r"C:\Users\USER\Desktop\지출내역서 관리앱\엑셀상세파일\입출금내역_" + datetime.now().strftime('%Y')+'.xlsx')
# 파일 없으면 포맷 만들기
except FileNotFoundError:
    # 파일 생성
    wb = Workbook()
    ws = wb.active
    # 12월까지 시트 만들기
    for l in range(1, 13):
        wb.create_sheet(f"{l}월")
       
        # 각 시트마다 헤더 만들기
        headers = ["날짜", "입금액", "출금액", "입금처", "출금처", "잔액"]
        ws = wb[f"{l}월"]
        ws.append(headers)
       
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15 # 날짜
        ws.column_dimensions['B'].width = 12 # 입금액
        ws.column_dimensions['C'].width = 12 # 출금액 
        ws.column_dimensions['D'].width = 35 # 입금처
        ws.column_dimensions['E'].width = 35 # 출금처
        ws.column_dimensions['F'].width = 15 # 잔액
      
        # 헤더 폰트 굵게
        for col in range(1, 7):
            ws.cell(1, col).font = Font(bold=True)
       
        # 초기날짜 먼저 입력 아랫작업을 안 하면 while문 실행이 안 됨
        ws['A2'] = dlt.strftime('%Y-%m-%d')
        # 초기잔액 계산식 입력
        dts = datetime(datetime.now().year, l, 1)
        dts -= timedelta(days=1)
        ws['F2'] = f"=B2-C2+'{dts.month}월'!F{dts.day+1}"
       
        # 초기날짜 입력했으니 1일 더하기
        dlt += timedelta(days=1)
      
        # A2라인에 썼으니 i도 2부터 시작
        i = 2
      
        # 각 날짜 입력
        while dlt.strftime('%d') != '01':
            i += 1
            ws[f"A{i}"] = dlt.strftime('%Y-%m-%d')
            dlt += timedelta(days=1)
            # 잔액 계산식 입력
            ws[f"F{i}"] = f"=B{i}-C{i}+F{i-1}"
       
        # 총합 계산
        ws[f"A{i+1}"] = "총합"
        ws[f"B{i+1}"] = f"=SUM(B1:B{i})"
        ws[f"C{i+1}"] = f"=SUM(C1:C{i})"
        
        #스타일 적용
        for row in ws['A2':f"F{i+1}"]:
            for cell in row:
                cell.style = style
        
        # 잔액 계산식 입력
            

    dty = datetime.now()
    dty -= timedelta(days=365)

    # 1년 전 파일에서 잔액 가져오기, 없으면 0으로 처리
    try:
        load_workbook(rf"C:\Users\USER\Desktop\지출내역서 관리앱\엑셀상세파일\입출금내역_{dty.strftime('%Y')}.xlsx")
        ws = wb[f"1월"]
        ws[f"F2"] = rf"='C:\Users\USER\Desktop\지출내역서 관리앱\엑셀상세파일" + f"\[입출금내역_{dty.strftime('%Y')}.xlsx]" + "12월'!$F$32"
        print (ws[f"F2"].value)
    except FileNotFoundError:
        ws = wb[f"1월"]
        with open(rf"C:\Users\USER\Desktop\지출내역서 관리앱\python_money-calculator\Set.txt", 'r', encoding='utf-8') as f:
                m = f.read()
                if m == None or m == "":
                    ws[f"F2"] = 0
                else:
                    ws[f"F2"] = int(m)
    #절대경로로 파일저장
    wb.save(r"C:\Users\USER\Desktop\지출내역서 관리앱\엑셀상세파일\입출금내역_" + datetime.now().strftime('%Y')+'.xlsx')

# 포맷이후 실행할 코드

def add(C, R):
    ws = wb[f"{datetime.now().month}월"]
    n = ws[f"B{int(datetime.now().strftime('%d'))+1}"].value
    a = ws[f"D{int(datetime.now().strftime('%d'))+1}"].value
    if n != None:
        n += C
        a = a + ", " + R
    else:
        n = C
        a = R
    ws[f"B{int(datetime.now().strftime('%d'))+1}"] = n
    ws[f"D{int(datetime.now().strftime('%d'))+1}"] = a

def minus(C, R):
    ws = wb[f"{datetime.now().month}월"]
    n = ws[f"C{int(datetime.now().strftime('%d'))+1}"].value
    a = ws[f"E{int(datetime.now().strftime('%d'))+1}"].value
    if n != None:
        n += C
        a = a + ", " + R
    else:
        n = C
        a = R
    ws[f"C{int(datetime.now().strftime('%d'))+1}"] = n
    ws[f"E{int(datetime.now().strftime('%d'))+1}"] = a

    wb.save(r"C:\Users\USER\Desktop\지출내역서 관리앱\엑셀상세파일\입출금내역_" + datetime.now().strftime('%Y')+'.xlsx')
